from sqlite3 import Row
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import re

#load existing spreadsheet
wb = load_workbook('testData.xlsx')

#create active worksheet 
ws = wb.active

#create new blank workbook
wbNew = Workbook()
sheet = wbNew.active 


#Grab column A
column_a = ws['A']
title = []
for cell in column_a:
    title.append(re.sub("[.]","", str(cell.value)))
    if ("Date of Possesion" in str(cell.value)):
        break

#append column A to new sheet
sheet.append(title)
#delete first 2 column of title 
sheet.delete_cols(1,2)

#Grab column B
column_b = ws['B']
#convert to list
column_b_list = list(column_b)
#slice list
column_b_list = column_b_list[2:]
residentDetails = []
count = 0

for cell in column_b_list:
    residentDetails.append(str(cell.value))
    count += 1
    if (count>=52):
        sheet.append(residentDetails)
        residentDetails =[]
        count=0

#clean the new sheet of all the columns that are not used
sheet.delete_cols(1,5)
sheet.delete_cols(16,50)


    # if ("None" in str(cell.value)):
    #     count += 1
    #     if count > 32:
    #         # if 32 consecutive None indicates new person
    #         sheet.append(residentDetails)
    #         residentDetails = []
    #         count = 0
    # else:
    #     count = 0

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
wbNew.save(r"C:\Users\Wei Liang\Desktop\Converted.xlsx")



# print (title)
    # if (cell.value is not None and cell.value !=''):
        # print(re.sub("[.]","", cell.value))